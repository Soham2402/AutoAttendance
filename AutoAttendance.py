from re import S
import cv2
import numpy as np
import pytesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
import os
from openpyxl import load_workbook, Workbook
import openpyxl
import os
import time

# converting all names into lowercase 
def lower(studentlist):
    for name in studentlist:
        studentlist[studentlist.index(name)] = name.lower()


#Finding img files and appending it into a list
def loopFiles():
    filelist = os.listdir(os.getcwd())
    imglist = []
    for i in filelist:
        if i.lower().endswith(('.png','.jpg','jpeg')):
                 imglist.append(i)
    return imglist

def cleanstring(namelist):
    formatlist = []
    for i in namelist:
        namestring = i.split("‘")
        namestring = "".join(namestring)
        formatlist.append(namestring)
    return formatlist

def checkpresent(studentlist):
    list1 = []
    present = []
    imglist = loopFiles()
    for image in imglist: 
        #reading names from image
        img = cv2.imread(image)
        text = pytesseract.image_to_string(img)
        for i in text.splitlines(): #READING THE IMAGE AND TURNING IT INTO A ROUGH LIST STRING 
            list1.append(i.lower())
    
    list1 = cleanstring(list1)

    #Checking names from student list and appending it into present list
    for studname in studentlist:
        for i in list1:
            seperatlist = i.split()
            for name in range(1,len(seperatlist)):
                if studname in seperatlist or studname == seperatlist[name-1] +" "+ seperatlist[name]:                    
                    present.append(studname.capitalize())
                    break 
    return present         


if __name__ == "__main__":

    #List of all students in the class
    studentlist = ["Gautham","Shreeju","Akshay","Anusaranya","Chandrasai",
                    "Aromal","Mihir","dona","Animesh","Ashwin","rahul",
                    "gayatri","shivam","gladys","eashwar","mahesh","Dineshkumar","Raj",
                    "Yash","Chandrakala","Ramyata","Harish","Soham","Abrial","Jivamani","Sayali","Tejas",
                    "Bhavana","Bhushan","Arun","Kartik","Pranesh","Ashreen","Raina","Mitesh arun","Shubhada","Ishita",
                    "Siddhesh","Nitin","Akash","Sen","Akshara","Pranit","Kevin","Adil","Anees","Devendra","Shoaib",
                    "Muthumari","Ajith","Santhni","Santhni","Harsh","Devika","Rishikesh","Rohit","Somdatt",
                    "Gauri","Brinda","Harikrishnan","Mahi","Pillai","Vishnu","Darshan","Amey","Tejal","Ankush","Roshani",
                    "Meenakshi","Sadiya","Aanam","Arman","Shalu","Divyesh","Shrinidhi","Varun","Huda","Vimal","Gautam",
                    "Harshit","Sneha","Shivani","Tejaswin","Sanjana","Sivakumar","Vamshi","Pratik","Mrunali","Rohit",
                    "Imran","Saikrishna","Hasnain","Kasim","Aftab","Aditya","Amirthavarshini","Hrithik","EASHWER","Nadar","pranav""Rutuja Ravindra","Rutuja Rane", "Rutuja Santosh","Shubham Vinod","Shubham Bhalerao",
                    "Nikhil sharadbabu","Nikhil srinivas","Pranav Salunke","Pranav Bindu"]

    

    #List of students with the same name
    wb = load_workbook("Attendance.xlsx")

    #WORKSHEETS
    lect = input("Enter the name of the class ").upper()

    ws = wb[lect]

    cell = input("Enter the cell: ")

    print("Please wait while we count attendance for you......")
    ##Turning all text to upper case as all elements in the recognised text will be upper
    lower(studentlist)


    Present = checkpresent(studentlist)
    print(Present)
    print(len(Present))
    

    now = time.strftime("%x") #date
    ws[f"{cell}1"].value = now



    #THIS PART MARKS THE STUDENT PRESENT
    for i in range(2,111):
        for j in Present:
            if ws[f"A{i}"].value == j:
                ws[f"{cell}{i}"].value = "P"


    wb.save("Attendance.xlsx")




    
    #Incase ub need to add students in a new worksheet

    # rows = 2
    # for i in studentlist:
    #     ws[f"A{rows}"].value = i.capitalize()
    #     rows = rows+1

    pres = checkpresent(studentlist)
    print(len(pres))
